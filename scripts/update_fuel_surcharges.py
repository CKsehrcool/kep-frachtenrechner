#!/usr/bin/env python3
"""
UPS Kraftstoffzuschlag Updater
===============================
Liest die aktuellen Kraftstoffzuschläge von der UPS-Website und aktualisiert
die Excel-Datei im KEP Frachtenrechner Repository.

Ausführung: python scripts/update_fuel_surcharges.py
"""

import re
import sys
import os
import openpyxl
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── Konfiguration ─────────────────────────────────────────────────────────────

EXCEL_FILENAME = "0493_Frachtenrechner_KEP_DATA.xlsx"
SHEET_NAME = "adds"
UPS_URL = "https://www.ups.com/de/de/support/shipping-support/shipping-costs-rates/fuel-surcharges"

# Tarif-Zuordnung: Welche Tarife bekommen den Standard-, welche den Express-Satz
STANDARD_TARIFFS = {
    "StandardSingleDE", "StandardSingleALL",
    "StandardMultiDE", "StandardMultiALL"
}
EXPRESS_TARIFFS = {
    "ExpressSaverALL_env", "ExpressSaverALL_doc", "ExpressSaverALL_pkg",
    "ExpressALL_env",      "ExpressALL_doc",      "ExpressALL_pkg",
    "ExpressNoon_env",     "ExpressNoon_pkg"
}

# Plausibilitätsbereiche (als Dezimalzahl)
STANDARD_RATE_MIN = 0.10   # 10 %
STANDARD_RATE_MAX = 0.50   # 50 %
EXPRESS_RATE_MIN  = 0.20   # 20 %
EXPRESS_RATE_MAX  = 0.70   # 70 %

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

def parse_percent(text: str) -> float | None:
    """
    Wandelt einen deutschen/englischen Prozentwert-String in eine Dezimalzahl um.
    z.B. '25,50 %' → 0.255  |  '43.5%' → 0.435
    """
    text = text.strip().replace("\xa0", "").replace(" ", "").replace("%", "")
    text = text.replace(",", ".")
    try:
        return round(float(text) / 100, 6)
    except ValueError:
        return None


def find_rates_in_text(text: str) -> tuple[float | None, float | None]:
    """
    Strategien zum Finden der Zuschlagssätze im Seitentext.
    Gibt (standard_rate, express_rate) zurück.
    """
    standard_rate = None
    express_rate  = None

    # Alle Prozentzahlen aus dem Text extrahieren
    pct_pattern = re.compile(r'(\d{1,3}[,\.]\d{1,2})\s*%')
    all_matches = pct_pattern.findall(text)
    print(f"  Alle gefundenen Prozentwerte: {all_matches}")

    # Volltext in Zeilen aufteilen für kontextsensitive Suche
    lines = text.splitlines()

    for i, line in enumerate(lines):
        line_lower = line.lower()

        # Kontext-Fenster: aktuelle + umliegende Zeilen
        context = " ".join(lines[max(0, i-2) : i+3]).lower()

        percents_in_context = pct_pattern.findall(context)
        if not percents_in_context:
            continue

        # Prüfe ob "Standard" im Kontext vorkommt
        if "standard" in context and "express" not in context:
            for p in percents_in_context:
                rate = parse_percent(p)
                if rate and STANDARD_RATE_MIN <= rate <= STANDARD_RATE_MAX:
                    if standard_rate is None:
                        standard_rate = rate
                        print(f"  → Standard-Satz gefunden: {rate:.4f} ({rate*100:.2f}%) in: '{line.strip()}'")

        # Prüfe ob "Express" im Kontext vorkommt
        if "express" in context:
            for p in percents_in_context:
                rate = parse_percent(p)
                if rate and EXPRESS_RATE_MIN <= rate <= EXPRESS_RATE_MAX:
                    if express_rate is None:
                        express_rate = rate
                        print(f"  → Express-Satz gefunden: {rate:.4f} ({rate*100:.2f}%) in: '{line.strip()}'")

    return standard_rate, express_rate


def find_rates_in_tables(page) -> tuple[float | None, float | None]:
    """
    Analysiert alle Tabellen auf der Seite nach Zuschlagssätzen.
    Gibt (standard_rate, express_rate) zurück.
    """
    standard_rate = None
    express_rate  = None

    tables = page.query_selector_all("table")
    print(f"  Tabellen gefunden: {len(tables)}")

    for t_idx, table in enumerate(tables):
        rows = table.query_selector_all("tr")
        header_map = {}     # Spaltenindex → Bezeichnung

        for r_idx, row in enumerate(rows):
            cells = row.query_selector_all("th, td")
            cell_texts = [c.inner_text().strip() for c in cells]
            if not cell_texts:
                continue

            row_lower = " ".join(cell_texts).lower()

            # Header-Zeile erkennen
            if r_idx == 0 or any(kw in row_lower for kw in ["standard", "express", "zuschlag", "surcharge"]):
                for ci, ct in enumerate(cell_texts):
                    ct_lower = ct.lower()
                    if "standard" in ct_lower:
                        header_map[ci] = "standard"
                    elif "express" in ct_lower:
                        header_map[ci] = "express"
                if header_map:
                    print(f"  Tabelle {t_idx}: Header-Map = {header_map}")
                continue

            # Datenzellen analysieren
            if header_map:
                pct_pattern = re.compile(r'(\d{1,3}[,\.]\d{1,2})\s*%')
                for ci, ct in enumerate(cell_texts):
                    matches = pct_pattern.findall(ct)
                    if not matches:
                        continue
                    rate = parse_percent(matches[0])
                    if rate is None:
                        continue
                    col_type = header_map.get(ci)
                    if col_type == "standard" and STANDARD_RATE_MIN <= rate <= STANDARD_RATE_MAX:
                        if standard_rate is None:
                            standard_rate = rate
                            print(f"  → Standard (Tabelle {t_idx}): {rate:.4f}")
                    elif col_type == "express" and EXPRESS_RATE_MIN <= rate <= EXPRESS_RATE_MAX:
                        if express_rate is None:
                            express_rate = rate
                            print(f"  → Express (Tabelle {t_idx}): {rate:.4f}")

    return standard_rate, express_rate


# ── Haupt-Scraping-Funktion ───────────────────────────────────────────────────

def scrape_ups_fuel_surcharges() -> tuple[float, float]:
    """
    Ruft die aktuellen UPS Kraftstoffzuschläge von der Website ab.
    Gibt (standard_rate, express_rate) als Dezimalzahlen zurück.
    """
    print(f"\n[1] Öffne UPS-Seite: {UPS_URL}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
        )
        context = browser.new_context(
            locale="de-DE",
            user_agent=(
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
        )
        page = context.new_page()
        screenshot_path = "ups_debug_screenshot.png"

        try:
            page.goto(UPS_URL, wait_until="networkidle", timeout=60_000)
        except Exception as nav_err:
            print(f"  Navigation-Fehler: {type(nav_err).__name__}: {nav_err}")

        # Cookie-Banner wegklicken (falls vorhanden)
        for selector in [
            "#onetrust-accept-btn-handler",
            "[id*='accept'][id*='cookie']",
            "[class*='accept-cookie']",
            "button[contains(@text,'Akzeptieren')]",
        ]:
            try:
                page.click(selector, timeout=2_000)
                page.wait_for_timeout(800)
                print("  Cookie-Banner akzeptiert")
                break
            except Exception:
                pass

        # Kurz warten, damit dynamische Inhalte laden
        page.wait_for_timeout(3_000)

        # Screenshot für Debugging speichern (auch bei Fehler)
        try:
            page.screenshot(path=screenshot_path, full_page=True)
            print(f"  Screenshot gespeichert: {screenshot_path}")
        except Exception as ss_err:
            print(f"  Screenshot fehlgeschlagen: {ss_err}")

        # ── Strategie 1: Tabellen ────────────────────────────────────────────
        print("\n[2] Strategie 1: Tabellen-Analyse")
        standard_rate, express_rate = find_rates_in_tables(page)

        # ── Strategie 2: Volltext ────────────────────────────────────────────
        if standard_rate is None or express_rate is None:
            print("\n[3] Strategie 2: Volltext-Analyse")
            body_text = page.inner_text("body")

            # Debug-Ausgabe (erste 3000 Zeichen)
            print("  --- Seiteninhalt (Auszug) ---")
            print(body_text[:3000])
            print("  ---")

            s, e = find_rates_in_text(body_text)
            standard_rate = standard_rate or s
            express_rate  = express_rate  or e

        browser.close()

    # ── Ergebnis-Validierung ─────────────────────────────────────────────────
    if standard_rate is None or express_rate is None:
        raise ValueError(
            f"Kraftstoffzuschläge konnten nicht extrahiert werden! "
            f"Standard={standard_rate}, Express={express_rate}\n"
            "Bitte Screenshot 'ups_debug_screenshot.png' prüfen."
        )

    if not (STANDARD_RATE_MIN <= standard_rate <= STANDARD_RATE_MAX):
        raise ValueError(f"Standard-Satz außerhalb des Plausibilitätsbereichs: {standard_rate}")
    if not (EXPRESS_RATE_MIN <= express_rate <= EXPRESS_RATE_MAX):
        raise ValueError(f"Express-Satz außerhalb des Plausibilitätsbereichs: {express_rate}")

    return standard_rate, express_rate


# ── Excel-Update ──────────────────────────────────────────────────────────────

def update_excel(
    excel_path: Path,
    standard_rate: float,
    express_rate: float
) -> tuple[bool, dict]:
    """
    Aktualisiert die Fuel Surcharge-Werte in der Excel-Datei.
    Gibt (changed, old_values) zurück.
    """
    wb = openpyxl.load_workbook(excel_path)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt '{SHEET_NAME}' nicht in {excel_path} gefunden!")

    ws    = wb[SHEET_NAME]
    changed    = False
    old_values = {}

    for row in ws.iter_rows(min_row=2):
        tarif = row[0].value
        if tarif is None:
            continue

        old_val  = row[1].value
        new_val  = None

        if tarif in STANDARD_TARIFFS:
            new_val = standard_rate
        elif tarif in EXPRESS_TARIFFS:
            new_val = express_rate

        if new_val is not None and old_val != new_val:
            old_values[tarif] = old_val
            row[1].value = new_val
            changed = True

    if changed:
        wb.save(excel_path)

    return changed, old_values


# ── Einstiegspunkt ────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  UPS Kraftstoffzuschlag Updater")
    print("=" * 60)

    # Excel-Datei suchen
    excel_path = Path(EXCEL_FILENAME)
    if not excel_path.exists():
        # Auch im Repo-Root suchen
        candidates = list(Path(".").glob("*.xlsx"))
        print(f"  {EXCEL_FILENAME} nicht gefunden. Verfügbare xlsx: {candidates}")
        sys.exit(1)

    print(f"\nExcel-Datei: {excel_path.resolve()}")

    # Aktuelle Werte aus Excel lesen (für Vergleich)
    wb_check = openpyxl.load_workbook(excel_path)
    ws_check = wb_check[SHEET_NAME]
    current_standard = next(
        (r[1].value for r in ws_check.iter_rows(min_row=2) if r[0].value in STANDARD_TARIFFS), None
    )
    current_express = next(
        (r[1].value for r in ws_check.iter_rows(min_row=2) if r[0].value in EXPRESS_TARIFFS), None
    )
    print(f"Aktuelle Werte → Standard: {current_standard}, Express: {current_express}")

    # Neue Werte von UPS holen
    try:
        new_standard, new_express = scrape_ups_fuel_surcharges()
    except Exception as e:
        print(f"\n❌ Fehler beim Scraping: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"\n✅ Neue Werte → Standard: {new_standard:.4f} ({new_standard*100:.2f}%)"
          f"  |  Express: {new_express:.4f} ({new_express*100:.2f}%)")

    # Excel aktualisieren
    changed, old_vals = update_excel(excel_path, new_standard, new_express)

    if changed:
        print(f"\n📝 Excel aktualisiert ({len(old_vals)} Zellen geändert):")
        for tarif, old in old_vals.items():
            new = new_standard if tarif in STANDARD_TARIFFS else new_express
            print(f"   {tarif}: {old} → {new}")
        # Diesen Output nutzt der GitHub Action für die Commit-Message
        print(f"\n::set-output name=changed::true")
        print(f"::set-output name=standard_rate::{new_standard}")
        print(f"::set-output name=express_rate::{new_express}")
    else:
        print("\n✅ Keine Änderungen – Werte sind bereits aktuell.")
        print("::set-output name=changed::false")


if __name__ == "__main__":
    main()
